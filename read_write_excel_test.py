#In this script, I simply want to read an file from the ABI7500 and re-write it as a new file.

#this block of text is how I would read a txt file
#f = open("E:/OneDrive/Research/Rockwell_Lab/Python/092017_Influenza_Diet_Secondary_IFNg_Blimp-1_RF.txt","r")
#print((f.read())
#f.close()

"""#I think I need pandas to read excel?
import pandas as pd
xl = pd.ExcelFile("E:/OneDrive/Research/Rockwell_Lab/Python/092017_Influenza_Diet_Secondary_IFNg_Blimp-1_RF.xls")
#Can add sheet names, but my file has none.  Will test this feature out later.
xl.sheet_names
    [u'Sheet1']

df = xl.parse("Sheet1")
df.head()"""

#For files with only one sheet_names
import pandas as pd
df = pd.read_excel("E:/OneDrive/Research/Rockwell_Lab/Python/092017_Influenza_Diet_Secondary_IFNg_Blimp-1_RF.xls")
print(df.head())

"""#Alternative method
from openpyxl import load_workbook
wb2 = load_workbook('E:/OneDrive/Research/Rockwell_Lab/Python/092017_Influenza_Diet_Secondary_IFNg_Blimp-1_RF.xls')
print(wb2.get_sheet_names())
['Sheet2', 'New Title', 'Sheet1']

worksheet1 = wb2['Sheet1'] # one way to load a worksheet
worksheet2 = wb2.get_sheet_by_name('Sheet2') # another way to load a worksheet
print(worksheet1['D18'].value)
for row in worksheet1.iter_rows():
    print(row[0].value())"""

#new_file = open("E:/OneDrive/Research/Rockwell_Lab/Python/rewritten_data.xls","w")
#new_file.write(str(f.read()))
#new_file.close()
#f.close()